import os
import json
import inspect
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

from ascend_fd.pkg.diag import fault_entity
from ascend_fd.utils.tool import CONF_PATH
from ascend_fd.pkg.diag.root_cluster import fault_description

PRE_NET = "NET_"
PRE_NODE = "NODE_"
REDUNDANT_WIDTH = 5
ASCEND_KG_CONFIG_FILE = os.path.join(CONF_PATH, "kg-config.json")
TXT_FONT = Font(name='宋体', size=11, bold=False)
HEADER_FONT = Font(name='宋体', size=11, bold=True)
TXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BORDER = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
HEADER_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


def get_json_event():
    """
    Obtains fault events defined in the ascend-kg-config.json file.
    :return: fault events
    """
    with open(ASCEND_KG_CONFIG_FILE, encoding='utf-8') as f:
        data = json.load(f)
    kg_data = data.get("knowledge-repository")
    return kg_data


def get_py_event(module):
    """
    Obtain the fault events defined in the python file.
    :param module: module
    :return: fault event instances
    """
    instances = {}
    members = inspect.getmembers(module, inspect.isclass)
    classes = [obj for name, obj in members if obj.__module__ == module.__name__]
    if len(classes) != 1:
        return instances
    for name, obj in module.__dict__.items():
        if inspect.isclass(obj):
            # skip class definitions
            continue
        if isinstance(obj, classes[0]):
            instances[name] = obj
    return instances


class EventSummary:
    def __init__(self):
        """
        Defining workbook
        """
        self.workbook = Workbook()

    @staticmethod
    def set_header_format(worksheet, columns):
        """
        Format header
        :param worksheet: excel worksheet
        :param columns: names of each column
        """
        worksheet.append(columns)
        for cell in worksheet[1]:
            cell.font, cell.alignment, cell.border, cell.fill = HEADER_FONT, HEADER_ALIGNMENT, BORDER, HEADER_FILL

    @staticmethod
    def set_content_format(worksheet, len_cell_value):
        """
        Format content
        :param worksheet: excel worksheet
        :param len_cell_value: width of each column of cells
        """
        for index, col in enumerate(worksheet.columns):
            column_letter = col[0].column_letter
            # adaptive column width
            worksheet.column_dimensions[column_letter].width = len_cell_value[index] + REDUNDANT_WIDTH
            for cell in col[1:]:  # need to exclude headers
                cell.font, cell.alignment, cell.border = TXT_FONT, TXT_ALIGNMENT, BORDER

    def _root_cause_device_analysis(self):
        """
        Root cause device analysis
        """
        worksheet = self.workbook.active
        worksheet.title = '根因设备分析故障事件'

        columns = ["一级分类(class)", "二级分类(component)", "三级分类(module)", "故障事件编号", '故障事件名称',
                   "故障事件原因", "处理建议"]

        self.set_header_format(worksheet, columns)

        json_data = get_json_event()
        len_cell_value = [len(txt) for txt in columns]
        for code, value in json_data.items():
            attr_data = value.get("attribute")
            class_data = attr_data.get("class")
            component_data = attr_data.get("component")
            module_data = attr_data.get("module")
            cause_zh_data = attr_data.get("cause_zh")
            description_zh_data = attr_data.get("description_zh")
            suggestion_zh = attr_data.get("suggestion_zh")
            all_data = [class_data, component_data, module_data, code, cause_zh_data, description_zh_data,
                        str(suggestion_zh)]

            worksheet.append(all_data)
            # record the maximum length of data in each column
            for index, data in enumerate(all_data):
                if len(data) > len_cell_value[index]:
                    len_cell_value[index] = len(data)

        self.set_content_format(worksheet, len_cell_value)

    def _network_congestion_and_device_resource_analysis(self):
        """
        Network congestion analysis and device resource analysis
        """
        columns = ["故障事件编号", "故障事件名称", "故障事件原因"]
        worksheet_net = self.workbook.create_sheet('网络拥塞分析故障事件')
        self.set_header_format(worksheet_net, columns)

        worksheet_node = self.workbook.create_sheet('设备资源分析故障事件')
        self.set_header_format(worksheet_node, columns)

        instances = get_py_event(fault_entity)
        len_cell_value = [len(txt) for txt in columns]
        for _, obj in instances.items():
            attribute = obj.attribute
            all_data = [attribute.get("code"), attribute.get("cause_zh"), attribute.get("description_zh")]
            if PRE_NET in obj.code:
                worksheet_net.append(all_data)
            if PRE_NODE in obj.code:
                worksheet_node.append(all_data)
            # record the maximum length of data in each column
            for index, data in enumerate(all_data):
                if len(data) > len_cell_value[index]:
                    len_cell_value[index] = len(data)

        self.set_content_format(worksheet_net, len_cell_value)
        self.set_content_format(worksheet_node, len_cell_value)

    def _root_cause_node_analysis(self):
        """
        Root cause node analysis
        """
        worksheet = self.workbook.create_sheet('根因节点分析故障场景')
        columns = ["场景编号", "场景名称", "场景描述"]
        self.set_header_format(worksheet, columns)

        len_cell_value = [len(txt) for txt in columns]
        instances = get_py_event(fault_description)
        for name, obj in instances.items():
            all_data = [obj.code, name, obj.string]
            worksheet.append(all_data)

            # record the maximum length of data in each column
            for index, data in enumerate(all_data):
                len_data = len(str(data))
                if len_data > len_cell_value[index]:
                    len_cell_value[index] = len_data

        self.set_content_format(worksheet, len_cell_value)

    def create_sheet(self, file_name):
        """
        Creating an excel table
        """
        self._root_cause_device_analysis()
        self._network_congestion_and_device_resource_analysis()
        self._root_cause_node_analysis()
        file_path = os.path.join(os.getcwd(), file_name)
        self.workbook.save(file_path)
        print(f"fault type excel created successfully, path: {file_path}")


if __name__ == "__main__":
    """
    生成资料故障诊断类型附件。
    1、表中code列是为了方便资料查找，归档时需提醒资料同事删除
    2、安装三方依赖：
        pip3 install ply
        pip3 install openpyxl
    """
    event_summary = EventSummary()
    event_summary.create_sheet("MindCluster 26.0.0 故障诊断类型.xlsx")
